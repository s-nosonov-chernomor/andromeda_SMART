# app/api/routes/settings.py
from __future__ import annotations

from fastapi import APIRouter, UploadFile, File, HTTPException, Request
from fastapi.responses import StreamingResponse, JSONResponse
from pydantic import BaseModel, Field
from typing import Any, Dict, List, Optional
from io import BytesIO
import yaml, copy, io
from copy import deepcopy
from openpyxl import Workbook, load_workbook

from app.core.config import settings
from app.services.hot_reload import start_lines, stop_lines, hot_reload_lines
from app.core.validate_cfg import validate_cfg


# ─────────────────────────────────────────────────────────────────────────────
# helpers
# ─────────────────────────────────────────────────────────────────────────────
def _normalize_lines_for_yaml(cfg: Dict[str, Any]) -> Dict[str, Any]:
    """
    Приводит lines к аккуратному виду:
      - порядок ключей фиксирован
      - для tcp не сохраняем serial-поля
      - для serial не сохраняем tcp-поля
    """
    lines = cfg.get("lines") or []
    new_lines: List[Dict[str, Any]] = []

    for ln in lines:
        t = (ln.get("transport") or "serial").strip().lower()
        nodes = ln.get("nodes") or []

        if t == "tcp":
            clean = {
                "name": ln.get("name", ""),
                "transport": "tcp",
                "host": ln.get("host", ""),
                "port": ln.get("port", 502),
                "timeout": float(ln.get("timeout", 1.0)),
                "port_retry_backoff_s": int(ln.get("port_retry_backoff_s", 5)),
                "nodes": nodes,
            }
        else:
            clean = {
                "name": ln.get("name", ""),
                "transport": t,  # serial/rtu
                "device": ln.get("device", ""),
                "baudrate": int(ln.get("baudrate", 9600)),
                "timeout": float(ln.get("timeout", 0.1)),
                "parity": (ln.get("parity", "N") or "N"),
                "stopbits": int(ln.get("stopbits", 1)),
                "port_retry_backoff_s": int(ln.get("port_retry_backoff_s", 5)),
                "rs485_rts_toggle": bool(ln.get("rs485_rts_toggle", False)),
                "nodes": nodes,
            }

        new_lines.append(clean)

    cfg["lines"] = new_lines
    return cfg


def _to_str(v: Any) -> str:
    return "" if v is None else str(v)

def _parse_int(v: Any, default: Optional[int] = None) -> Optional[int]:
    if v is None or v == "":
        return default
    try:
        if isinstance(v, float):
            return int(v)
        return int(str(v).strip())
    except Exception:
        return default

def _parse_float_ru(v: Any, default: Optional[float] = None) -> Optional[float]:
    """
    Понимает:
      - float/int из openpyxl
      - строки с точкой или запятой: "0,001"
      - пустые -> default
    """
    if v is None or v == "":
        return default
    if isinstance(v, (int, float)):
        return float(v)
    try:
        s = str(v).strip().replace(" ", "").replace(",", ".")
        return float(s)
    except Exception:
        return default

def _migrate_param_ms_to_s(p: Dict[str, Any]) -> Dict[str, Any]:
    p = dict(p or {})
    if "publish_interval_ms" in p and "publish_interval_s" not in p:
        try:
            p["publish_interval_s"] = float(p["publish_interval_ms"]) / 1000.0
        except Exception:
            p["publish_interval_s"] = 0.0
        p.pop("publish_interval_ms", None)
    return p

def _cfg() -> Dict[str, Any]:
    cfg = settings.get_cfg()
    if not isinstance(cfg, dict):
        raise HTTPException(500, "Config is not loaded")
    return copy.deepcopy(cfg)

def _write_cfg(cfg: Dict[str, Any]) -> str:
    """Сохранить YAML через settings + вернуть имя backup-файла (или '')."""
    try:
        validate_cfg(cfg)  # ← проверяем ПЕРЕД записью
    except ValueError as e:
        # отдаём 400 в читаемом виде
        from fastapi import HTTPException
        raise HTTPException(400, f"Некорректный конфиг: {e}")

    backup_name = settings.save_yaml_config(cfg)
    return backup_name


def _find_line(cfg: Dict[str, Any], line_name: str) -> Optional[Dict[str, Any]]:
    for ln in cfg.get("lines", []) or []:
        if ln.get("name") == line_name:
            return ln
    return None

def _find_node(line: Dict[str, Any], unit_id: int) -> Optional[Dict[str, Any]]:
    for nd in line.get("nodes", []) or []:
        if int(nd.get("unit_id")) == int(unit_id):
            return nd
    return None

# ─────────────────────────────────────────────────────────────────────────────
# схемы
# ─────────────────────────────────────────────────────────────────────────────

class BatchRead(BaseModel):
    enabled: Optional[bool] = None
    max_bits: Optional[int] = None
    max_registers: Optional[int] = None

class Polling(BaseModel):
    interval_ms: Optional[int] = None
    jitter_ms: Optional[int] = None
    backoff_ms: Optional[int] = None
    max_errors_before_backoff: Optional[int] = None
    port_retry_backoff_s: Optional[int] = None
    batch_read: Optional[BatchRead] = None

class Mqtt(BaseModel):
    host: Optional[str] = None
    port: Optional[int] = None
    base_topic: Optional[str] = None
    qos: Optional[int] = None
    retain: Optional[bool] = None
    client_id: Optional[str] = None

class Debug(BaseModel):
    enabled: Optional[bool] = None
    log_reads: Optional[bool] = None
    summary_every_s: Optional[int] = None

class History(BaseModel):
    max_rows: Optional[int] = None
    ttl_days: Optional[int] = None
    cleanup_every: Optional[int] = None

class GeneralDTO(BaseModel):
    mqtt: Mqtt
    polling: Polling
    history: History
    debug: Debug

class ParamDTO(BaseModel):
    name: str
    register_type: str
    address: int
    scale: Optional[float] = 1.0
    mode: str = "r"
    publish_mode: Optional[str] = "on_change"
    publish_interval_s: Optional[float] = 0
    topic: Optional[str] = None
    step: Optional[float] = None
    hysteresis: Optional[float] = None
    # новые для multi-register
    words: Optional[int] = 1
    data_type: Optional[str] = "u16"       # u16|s16|u32|s32|f32
    word_order: Optional[str] = "AB"       # AB|BA

class AddParamDTO(BaseModel):
    line: str
    unit_id: int
    object: str
    # если узла нет — создадим; можно передать его номер
    num_object: Optional[int] = 1
    param: ParamDTO

class UpdateParamDTO(BaseModel):
    line: str
    unit_id: int
    name: str
    updates: ParamDTO

class LineDTO(BaseModel):
    name: str
    device: Optional[str] = ""
    baudrate: Optional[int] = 9600
    timeout: Optional[float] = 0.1
    parity: Optional[str] = "N"           # N/E/O
    stopbits: Optional[int] = 1           # 1/2
    port_retry_backoff_s: Optional[int] = 5
    rs485_rts_toggle: Optional[bool] = False
    nodes: Optional[List[Dict[str, Any]]] = None  # узлы с unit_id/object/num_object/params...

class LinesFullDTO(BaseModel):
    lines: List[LineDTO]

router = APIRouter(prefix="/api/settings", tags=["settings"])

# ─────────────────────────────────────────────────────────────────────────────
# enums (для UI)
# ─────────────────────────────────────────────────────────────────────────────

@router.get("/enums")
def get_enums():
    return {
        "register_types": ["coil", "discrete", "holding", "input"],
        "param_modes": ["r", "rw"],
        "publish_modes": ["on_change", "interval", "on_change_and_interval"],
        "parity": ["N", "E", "O"],
        "stopbits": [1, 2],
        # новые enum-ы для UI параметров
        "data_types": ["u16", "s16", "u32", "s32", "u64", "s64", "f32"],
        "word_orders": ["AB", "BA", "ABCD", "DCBA", "BADC", "CDAB"],
    }

# ─────────────────────────────────────────────────────────────────────────────
# Общие
# ─────────────────────────────────────────────────────────────────────────────

@router.get("/general")
def get_general():
    cfg = _cfg()
    # отдаем всё, что сейчас есть в YAML (с дефолтами где нужно)
    return {
        "mqtt": cfg.get("mqtt", {}),
        "db": cfg.get("db", {"url": "sqlite:///./data/data.db"}),
        "history": cfg.get("history", {}),
        "debug": cfg.get("debug", {}),
        "serial": cfg.get("serial", {"echo": False}),
        "addressing": cfg.get("addressing", {"normalize": True}),
        "backups": cfg.get("backups", {"dir": "./data/backups", "keep": 10}),
        "service": cfg.get("service", {"unit": "agent.service", "restart_cmd": ""}),
        "andromeda": cfg.get("andromeda", {"restart_cmd": "/usr/local/bin/restart-andromeda.sh"}),
        "current": cfg.get("current", {"touch_read_every_s": 3}),
        "polling": cfg.get("polling", {}),
    }

@router.put("/general")
def put_general(body: Dict[str, Any]):
    # требуем ключевые секции
    for k in ("mqtt", "db", "polling", "history", "debug", "serial", "addressing", "backups", "service", "andromeda", "current"):
        if k not in body:
            raise HTTPException(400, f"Missing section in body: {k}")

    cfg = _cfg()
    cfg["mqtt"] = body["mqtt"] or {}
    cfg["db"] = body["db"] or {"url": "sqlite:///./data/data.db"}
    cfg["polling"] = body["polling"] or {}
    cfg["history"] = body["history"] or {}
    cfg["debug"] = body["debug"] or {}
    cfg["serial"] = body["serial"] or {"echo": False}
    cfg["addressing"] = {"normalize": bool((body.get("addressing") or {}).get("normalize", True))}
    cfg["backups"] = body["backups"] or {"dir": "./data/backups", "keep": 10}
    cfg["service"] = body["service"] or {"unit": "agent.service", "restart_cmd": ""}
    cfg["andromeda"] = body["andromeda"] or {"restart_cmd": "/usr/local/bin/restart-andromeda.sh"}
    cfg["current"] = body["current"] or {"touch_read_every_s": 3}

    backup = _write_cfg(cfg)
    return {"ok": True, "backup": backup}


# ─────────────────────────────────────────────────────────────────────────────
# ЛИНИИ целиком (редактирование портов/скоростей и т.п.)
# ─────────────────────────────────────────────────────────────────────────────

@router.get("/lines")
def get_lines():
    cfg = _cfg()
    out = []
    for ln in (cfg.get("lines") or []):
        ln2 = copy.deepcopy(ln)
        for nd in (ln2.get("nodes") or []):
            for p in (nd.get("params") or []):
                mp = _migrate_param_ms_to_s(p)
                p.clear(); p.update(mp)
        out.append(ln2)
    return out


@router.post("/line/add")
def add_line(payload: Dict[str, Any]):
    name = (payload.get("name") or "").strip()
    if not name:
        raise HTTPException(400, "Missing line name")
    cfg = _cfg()
    lines: List[Dict[str, Any]] = cfg.setdefault("lines", [])
    if any((ln.get("name") == name) for ln in lines):
        raise HTTPException(409, "Line already exists")

    # дефолты не затирают YAML, просто добавляем новую
    lines.append({
        "name": name,
        "device": "",
        "baudrate": 9600,
        "timeout": 0.1,
        "parity": "N",
        "stopbits": 1,
        "port_retry_backoff_s": 5,
        "rs485_rts_toggle": False,
        "nodes": []
    })
    cfg = _normalize_lines_for_yaml(cfg)
    backup = _write_cfg(cfg)
    return {"ok": True, "backup": backup}

@router.put("/line/update")
def update_line(body: Dict[str, Any]):
    body = body or {}
    name = body.get("name")
    if not name:
        raise HTTPException(400, "Missing line name")

    cfg = _cfg()
    lines: List[Dict[str, Any]] = cfg.get("lines", [])
    line = next((ln for ln in lines if ln.get("name") == name), None)
    if not line:
        raise HTTPException(404, "Line not found")

    # поддерживаем оба формата: {name, updates:{...}} ИЛИ {name, device,...}
    updates = body.get("updates") or {
        k: body[k] for k in (
            "transport",
            "device", "baudrate", "parity", "stopbits",
            "timeout", "port_retry_backoff_s", "rs485_rts_toggle",
            "host", "port"
        ) if k in body
    }

    for k, v in updates.items():
        if k in (
            "transport",
            "device","baudrate","parity","stopbits","timeout","port_retry_backoff_s","rs485_rts_toggle",
            "host","port"
        ):
            line[k] = v

    cfg = _normalize_lines_for_yaml(cfg)
    backup = _write_cfg(cfg)
    return {"ok": True, "backup": backup}

@router.delete("/line/delete")
def delete_line(body: Dict[str, Any]):
    name = (body.get("name") or "").strip()
    if not name:
        raise HTTPException(400, "Missing line name")

    cfg = _cfg()
    lines: List[Dict[str, Any]] = cfg.get("lines", [])
    idx = next((i for i, ln in enumerate(lines) if ln.get("name") == name), -1)
    if idx < 0:
        raise HTTPException(404, "Line not found")

    # удаляем линию целиком (вместе с nodes/params)
    lines.pop(idx)

    backup = _write_cfg(cfg)
    return {"ok": True, "backup": backup}

@router.post("/node/add")
def add_node(body: Dict[str, Any]):
    body = body or {}
    node_block = body.get("node")
    if isinstance(node_block, dict):
        line_name = body.get("line")
        unit_id   = node_block.get("unit_id")
        object_   = (node_block.get("object") or "").strip()
        num_obj   = node_block.get("num_object", None)
    else:
        line_name = body.get("line")
        unit_id   = body.get("unit_id")
        object_   = (body.get("object") or "").strip()
        num_obj   = body.get("num_object", None)


    if not line_name or unit_id is None or not object_:
        raise HTTPException(400, "Bad payload")

    cfg = _cfg()
    lines: List[Dict[str, Any]] = cfg.get("lines", [])
    line = next((ln for ln in lines if ln.get("name") == line_name), None)
    if not line:
        raise HTTPException(404, "Line not found")

    nodes: List[Dict[str, Any]] = line.setdefault("nodes", [])
    if any(int(nd.get("unit_id", -1)) == int(unit_id) for nd in nodes):
        raise HTTPException(409, "Node with this unit_id already exists")

    node = {"unit_id": int(unit_id), "object": object_, "params": []}
    if num_obj is not None:
        node["num_object"] = int(num_obj)
    nodes.append(node)

    backup = _write_cfg(cfg)
    return {"ok": True, "backup": backup}

@router.put("/node/update")
def update_node(body: Dict[str, Any]):
    """
    body = {
      "line": "line1",
      "old_unit_id": 1,
      "updates": {"unit_id": 2, "object": "new_obj", "num_object": 5}
    }
    """
    body = body or {}
    line_name = body.get("line")
    old_unit_id = body.get("old_unit_id")
    updates = body.get("updates") or {}
    if not line_name or old_unit_id is None:
        raise HTTPException(400, "Bad payload (need line, old_unit_id)")

    cfg = _cfg()
    lines: List[Dict[str, Any]] = cfg.get("lines", [])
    line = next((ln for ln in lines if ln.get("name") == line_name), None)
    if not line:
        raise HTTPException(404, "Line not found")

    nodes: List[Dict[str, Any]] = line.setdefault("nodes", [])
    node = next((nd for nd in nodes if int(nd.get("unit_id", -1)) == int(old_unit_id)), None)
    if not node:
        raise HTTPException(404, "Node not found")

    if "unit_id" in updates:
        new_uid = int(updates["unit_id"])
        if new_uid != int(old_unit_id) and any(int(nd.get("unit_id", -1)) == new_uid for nd in nodes):
            raise HTTPException(409, "Another node with this unit_id already exists")
        node["unit_id"] = new_uid

    if "object" in updates:
        node["object"] = updates["object"]

    if "num_object" in updates:
        node["num_object"] = updates["num_object"]

    backup = _write_cfg(cfg)
    return {"ok": True, "backup": backup}

@router.delete("/node/delete")
def delete_node(body: Dict[str, Any]):
    line_name = body.get("line")
    unit_id   = body.get("unit_id")
    if not line_name or unit_id is None:
        raise HTTPException(400, "Bad payload")

    cfg = _cfg()
    lines: List[Dict[str, Any]] = cfg.get("lines", [])
    line = next((ln for ln in lines if ln.get("name") == line_name), None)
    if not line:
        raise HTTPException(404, "Line not found")

    nodes: List[Dict[str, Any]] = line.get("nodes", [])
    idx = next((i for i, nd in enumerate(nodes) if int(nd.get("unit_id", -1)) == int(unit_id)), -1)
    if idx < 0:
        raise HTTPException(404, "Node not found")

    nodes.pop(idx)
    backup = _write_cfg(cfg)
    return {"ok": True, "backup": backup}


# ─────────────────────────────────────────────────────────────────────────────
# Параметры (добавить/изменить/удалить) — бережное обновление
# ─────────────────────────────────────────────────────────────────────────────

@router.post("/param/add")
def add_param(payload: Dict[str, Any]):
    line_name = payload.get("line")
    unit_id   = payload.get("unit_id")
    object_   = payload.get("object")
    param     = payload.get("param")
    param = _migrate_param_ms_to_s(param)
    if not line_name or unit_id is None or not object_ or not isinstance(param, dict):
        raise HTTPException(400, "Bad payload")

    cfg = _cfg()
    lines: List[Dict[str, Any]] = cfg.setdefault("lines", [])
    line = next((ln for ln in lines if ln.get("name") == line_name), None)
    if not line:
        raise HTTPException(404, "Line not found")

    nodes: List[Dict[str, Any]] = line.setdefault("nodes", [])
    node  = next((nd for nd in nodes if int(nd.get("unit_id", -1)) == int(unit_id)), None)
    if not node:
        # создадим узел на лету, чтобы не падать
        node = {"unit_id": int(unit_id), "object": object_, "params": []}
        nodes.append(node)

    params: List[Dict[str, Any]] = node.setdefault("params", [])
    existing = next((p for p in params if p.get("name") == param.get("name")), None)
    if existing:
        params.remove(existing)
    params.append(param)

    backup = _write_cfg(cfg)
    return {"ok": True, "backup": backup}

@router.put("/param/update")
def update_param(body: Dict[str, Any]):
    line_name = body.get("line")
    unit_id   = body.get("unit_id")
    name      = body.get("name")
    updates   = body.get("updates")
    updates = _migrate_param_ms_to_s(updates)
    if not line_name or unit_id is None or not name or not isinstance(updates, dict):
        raise HTTPException(400, "Bad payload")

    cfg = _cfg()
    lines: List[Dict[str, Any]] = cfg.get("lines", [])
    line = next((ln for ln in lines if ln.get("name") == line_name), None)
    if not line:
        raise HTTPException(404, "Line not found")

    nodes: List[Dict[str, Any]] = line.get("nodes", [])
    node  = next((nd for nd in nodes if int(nd.get("unit_id", -1)) == int(unit_id)), None)
    if not node:
        raise HTTPException(404, "Node not found")

    params: List[Dict[str, Any]] = node.get("params", [])
    idx    = next((i for i, p in enumerate(params) if p.get("name") == name), -1)
    if idx < 0:
        raise HTTPException(404, "Param not found")

    # аккуратно поддержим rename
    current = params[idx].copy()
    newp = current.copy()
    newp.update(updates or {})

    new_name = newp.get("name", name)
    if new_name != name and any(p.get("name") == new_name for p in params):
        raise HTTPException(409, "Param with new name already exists")

    params[idx] = newp
    backup = _write_cfg(cfg)
    return {"ok": True, "backup": backup}

@router.delete("/param/delete")
def delete_param(body: Dict[str, Any]):
    line_name = body.get("line")
    unit_id   = body.get("unit_id")
    name      = body.get("name")
    if not line_name or unit_id is None or not name:
        raise HTTPException(400, "Bad payload")

    cfg = _cfg()
    lines: List[Dict[str, Any]] = cfg.get("lines", [])
    line = next((ln for ln in lines if ln.get("name") == line_name), None)
    if not line:
        raise HTTPException(404, "Line not found")

    nodes: List[Dict[str, Any]] = line.get("nodes", [])
    node  = next((nd for nd in nodes if int(nd.get("unit_id", -1)) == int(unit_id)), None)
    if not node:
        raise HTTPException(404, "Node not found")

    params: List[Dict[str, Any]] = node.get("params", [])
    idx    = next((i for i, p in enumerate(params) if p.get("name") == name), -1)
    if idx < 0:
        raise HTTPException(404, "Param not found")

    params.pop(idx)
    backup = _write_cfg(cfg)
    return {"ok": True, "backup": backup}

@router.get("/params/export")
def export_params_xlsx():
    cfg = settings.get_cfg()
    wb = Workbook()
    ws = wb.active
    ws.title = "params"

    headers = [
        # line key + line settings
        "Линия",
        "Transport",
        "Device",
        "Host",
        "Port",
        "Baudrate",
        "Parity",
        "Stopbits",
        "Timeout, s",
        "port_retry_backoff_s",
        "rs485_rts_toggle",

        # node key + node settings
        "Unit",
        "Object",
        "Num_object",

        # param
        "Параметр",
        "Тип",
        "Адрес",
        "Words",
        "DataType",
        "WordOrder",
        "Scale",
        "Mode",
        "Publish",
        "Interval, s",
        "Topic",
        "Step",
        "Hysteresis",
    ]
    ws.append(headers)

    # (не обязательно) ширины колонок чуть удобнее
    widths = {
        "A": 14, "B": 10, "C": 16, "D": 16, "E": 8,
        "F": 10, "G": 8, "H": 9, "I": 10, "J": 16, "K": 14,
        "L": 8, "M": 20, "N": 12,
        "O": 18, "P": 10, "Q": 10, "R": 7, "S": 10, "T": 10,
        "U": 16, "V": 8, "W": 14, "X": 12, "Y": 20, "Z": 10, "AA": 12
    }
    for col, w in widths.items():
        ws.column_dimensions[col].width = w

    for ln in cfg.get("lines", []) or []:
        line_name = ln.get("name", "")
        transport = (ln.get("transport") or "serial").strip().lower()

        prb = ln.get("port_retry_backoff_s", 5)

        # defaults as empty (so we never "invent" data)
        device = ""
        baudrate = ""
        parity = ""
        stopbits = ""
        host = ""
        port = ""
        timeout = ""

        rts = ""  # rs485_rts_toggle

        if transport == "tcp":
            host = ln.get("host", "") or ""
            port = ln.get("port", "") or ""
            timeout = ln.get("timeout", "") or ""
            # rs485 для tcp не имеет смысла
        else:
            device = ln.get("device", "") or ""
            baudrate = ln.get("baudrate", "") or ""
            parity = ln.get("parity", "") or ""
            stopbits = ln.get("stopbits", "") or ""
            timeout = ln.get("timeout", "") or ""
            rts = ln.get("rs485_rts_toggle", "")

        for nd in ln.get("nodes", []) or []:
            unit = nd.get("unit_id", "")
            obj = nd.get("object", "")
            num_obj = nd.get("num_object", "")

            for p in nd.get("params", []) or []:
                mp = _migrate_param_ms_to_s(p)

                ws.append([
                    line_name,
                    transport,
                    device,
                    host,
                    port,
                    baudrate,
                    parity,
                    stopbits,
                    timeout,
                    prb,
                    rts,

                    unit,
                    obj,
                    num_obj,

                    mp.get("name", ""),
                    mp.get("register_type", ""),
                    mp.get("address", ""),
                    mp.get("words", 1),
                    mp.get("data_type", "u16"),
                    mp.get("word_order", "AB"),
                    mp.get("scale", 1.0),
                    mp.get("mode", "r"),
                    mp.get("publish_mode", "on_change"),
                    mp.get("publish_interval_s", 0),
                    mp.get("topic") or "",
                    mp.get("step", None),
                    mp.get("hysteresis", None),
                ])

    buf = io.BytesIO()
    wb.save(buf)
    buf.seek(0)
    return StreamingResponse(
        buf,
        media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        headers={"Content-Disposition": 'attachment; filename="params.xlsx"'},
    )

@router.post("/params/import")
async def import_params_xlsx(file: UploadFile = File(...)):
    content = await file.read()
    try:
        wb = load_workbook(io.BytesIO(content), data_only=True)
    except Exception as e:
        raise HTTPException(400, f"XLSX parse error: {e}")

    ws = wb["params"] if "params" in wb.sheetnames else wb.active
    headers_row = [str(c.value or "").strip() for c in ws[1]]
    idx = {name: i for i, name in enumerate(headers_row)}

    required = ["Линия", "Unit", "Параметр", "Тип", "Адрес"]
    missing = [h for h in required if h not in idx]
    if missing:
        raise HTTPException(400, f"В файле отсутствуют колонки: {', '.join(missing)}")

    def cell(row, key):
        if key not in idx:
            return None
        return row[idx[key]]

    cfg = _cfg()

    # Будем собирать новый список lines ТОЛЬКО из файла (как у тебя сейчас)
    new_lines: Dict[str, Dict[str, Any]] = {}
    nodes_by_key: Dict[tuple, Dict[str, Any]] = {}
    total_params = 0

    # флаги "первая строка уже задавала настройки"
    line_settings_taken: Dict[str, bool] = {}
    node_meta_taken: Dict[tuple, bool] = {}

    for row in ws.iter_rows(min_row=2, values_only=True):
        if not any(row):
            continue

        line_name = _to_str(cell(row, "Линия")).strip()
        if not line_name:
            continue

        unit_id = _parse_int(cell(row, "Unit"), None)
        pname = _to_str(cell(row, "Параметр")).strip()
        reg_type = _to_str(cell(row, "Тип")).strip()
        address = _parse_int(cell(row, "Адрес"), None)

        if unit_id is None or not pname or not reg_type or address is None:
            continue

        # ---------- ensure line ----------
        line = new_lines.get(line_name)
        if not line:
            # если была такая линия в текущем YAML — возьмём дефолты из неё
            exist = _find_line(cfg, line_name)
            if exist:
                line = deepcopy(exist)
                line["nodes"] = []
            else:
                line = {
                    "name": line_name,
                    "transport": "serial",
                    "device": "",
                    "baudrate": 9600,
                    "timeout": 0.1,
                    "parity": "N",
                    "stopbits": 1,
                    "port_retry_backoff_s": 5,
                    "rs485_rts_toggle": False,
                    "nodes": [],
                }
            new_lines[line_name] = line

        # ---------- line settings (take first suitable row only) ----------
        # ---------- line settings (take first COMPLETE row only) ----------
        if not line_settings_taken.get(line_name, False):
            transport_cell = _to_str(cell(row, "Transport")).strip().lower()
            transport = transport_cell or (line.get("transport") or "serial")
            transport = (transport or "serial").strip().lower()

            if transport == "tcp":
                host_v = _to_str(cell(row, "Host")).strip()
                port_v = _parse_int(cell(row, "Port"), None)

                # "полная" строка для tcp
                if host_v and port_v is not None:
                    line["transport"] = "tcp"
                    line["host"] = host_v
                    line["port"] = int(port_v)

                    timeout = _parse_float_ru(cell(row, "Timeout, s"), None)
                    prb = _parse_int(cell(row, "port_retry_backoff_s"), None)
                    if timeout is not None:
                        line["timeout"] = float(timeout)
                    if prb is not None:
                        line["port_retry_backoff_s"] = int(prb)

                    line_settings_taken[line_name] = True

            else:
                dev_v = _to_str(cell(row, "Device")).strip()

                # "полная" строка для serial
                if dev_v:
                    line["transport"] = transport  # "serial"/"rtu"
                    line["device"] = dev_v

                    baudrate = _parse_int(cell(row, "Baudrate"), None)
                    parity = _to_str(cell(row, "Parity")).strip().upper()
                    stopbits = _parse_int(cell(row, "Stopbits"), None)
                    timeout = _parse_float_ru(cell(row, "Timeout, s"), None)
                    prb = _parse_int(cell(row, "port_retry_backoff_s"), None)
                    rts = cell(row, "rs485_rts_toggle")

                    if baudrate is not None:
                        line["baudrate"] = int(baudrate)
                    if parity in ("N", "E", "O"):
                        line["parity"] = parity
                    if stopbits in (1, 2):
                        line["stopbits"] = int(stopbits)
                    if timeout is not None:
                        line["timeout"] = float(timeout)
                    if prb is not None:
                        line["port_retry_backoff_s"] = int(prb)
                    if rts is not None and rts != "":
                        if isinstance(rts, bool):
                            line["rs485_rts_toggle"] = bool(rts)
                        else:
                            line["rs485_rts_toggle"] = str(rts).strip().lower() in ("1", "true", "да", "yes", "y")

                    line_settings_taken[line_name] = True


        # ---------- ensure node ----------
        nkey = (line_name, int(unit_id))
        node = nodes_by_key.get(nkey)
        if not node:
            # попытаемся взять object/num_object из текущего YAML как дефолт
            exist_ln = _find_line(cfg, line_name)
            exist_nd = _find_node(exist_ln, int(unit_id)) if exist_ln else None

            node = {"unit_id": int(unit_id), "params": []}
            if exist_nd:
                node["object"] = exist_nd.get("object", f"unit{unit_id}")
                if "num_object" in exist_nd:
                    node["num_object"] = exist_nd.get("num_object")
            else:
                node["object"] = f"unit{unit_id}"

            nodes_by_key[nkey] = node
            line.setdefault("nodes", []).append(node)

        # ---------- node meta (take first non-empty object/num_object for this unit) ----------
        if not node_meta_taken.get(nkey, False):
            obj = _to_str(cell(row, "Object")).strip()
            num_obj = _parse_int(cell(row, "Num_object"), None)

            if obj:
                node["object"] = obj
            if num_obj is not None:
                node["num_object"] = int(num_obj)

            if obj or num_obj is not None:
                node_meta_taken[nkey] = True

        # ---------- param fields (each row) ----------
        words = _parse_int(cell(row, "Words"), 1) or 1
        data_type = (_to_str(cell(row, "DataType")).strip() or "u16")
        word_order = (_to_str(cell(row, "WordOrder")).strip() or "AB")
        scale = _parse_float_ru(cell(row, "Scale"), 1.0) or 1.0
        mode = (_to_str(cell(row, "Mode")).strip() or "r")
        pmode = (_to_str(cell(row, "Publish")).strip() or "on_change")
        pint_s = _parse_float_ru(cell(row, "Interval, s"), 0.0) or 0.0
        topic = _to_str(cell(row, "Topic")).strip() or None
        step = _parse_float_ru(cell(row, "Step"), None)
        hyst = _parse_float_ru(cell(row, "Hysteresis"), None)

        param = {
            "name": pname,
            "register_type": reg_type,
            "address": int(address),
            "words": int(words),
            "data_type": data_type,
            "word_order": word_order,
            "scale": float(scale),
            "mode": mode,
            "publish_mode": pmode,
            "publish_interval_s": float(pint_s),
            "topic": topic,
        }
        if step is not None:
            param["step"] = float(step)
        if hyst is not None:
            param["hysteresis"] = float(hyst)

        param = _migrate_param_ms_to_s(param)

        # replace param with same name (чтобы импорт был идемпотентным)
        plist: List[Dict[str, Any]] = node.setdefault("params", [])
        prev = next((pp for pp in plist if pp.get("name") == pname), None)
        if prev:
            plist.remove(prev)
        plist.append(param)

        total_params += 1

    cfg["lines"] = list(new_lines.values())
    cfg = _normalize_lines_for_yaml(cfg)
    backup = _write_cfg(cfg)

    try:
        hot_reload_lines(cfg)
    except Exception as e:
        raise HTTPException(500, f"YAML сохранён (backup: {backup}), но перезапуск линий не удался: {e}")

    return {"ok": True, "backup": backup, "imported_params": total_params}


# ─────────────────────────────────────────────────────────────────────────────
# Файл на диск / перечитать / перезапустить линии
# ─────────────────────────────────────────────────────────────────────────────

@router.post("/read_disk")
def read_disk():
    settings.load_yaml_config()
    return {"ok": True}

@router.post("/save_disk")
def save_disk():
    backup = _write_cfg(_cfg())
    return {"ok": True, "backup": backup}

@router.post("/reload")
def reload_lines():
    hot_reload_lines(settings.get_cfg())
    return {"ok": True}
